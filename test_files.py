import pytest
import zipfile
from pypdf import PdfReader
import csv
import io
from openpyxl import load_workbook
from io import BytesIO

# Определите ваши пути к файлам
zip_filename = "tmp/archive.zip"
files_to_zip = ["tmp/test.pdf", "tmp/train.csv", "tmp/import_empl_xlsx.xlsx"]

def test_pdf_zip(create_zip_archive):
    create_zip_archive(zip_filename, files_to_zip)
    with zipfile.ZipFile(zip_filename, 'r') as zip_file:
        with zip_file.open("test.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page_text = reader.pages[0].extract_text()
            assert "Тестовый PDF файл" in page_text

def test_csv_zip(create_zip_archive):
    create_zip_archive(zip_filename, files_to_zip)
    with zipfile.ZipFile(zip_filename, 'r') as zip_file:
        with zip_file.open("train.csv") as csv_file:
            csv_content = csv_file.read()
            csv_text = csv_content.decode('utf-8')
            csv_file_like = io.StringIO(csv_text)
            reader = csv.reader(csv_file_like)
            header = next(reader)
            assert "Name" in header

def test_xlsx_zip(create_zip_archive):
    create_zip_archive(zip_filename, files_to_zip)
    with zipfile.ZipFile(zip_filename, 'r') as zip_file:
        with zip_file.open("import_empl_xlsx.xlsx") as xlsx_file:
            with BytesIO(xlsx_file.read()) as temp_file:
                workbook = load_workbook(temp_file)
                sheet = workbook.active
                text = sheet.cell(row=2, column=2).value
                assert "Иванова" in text
